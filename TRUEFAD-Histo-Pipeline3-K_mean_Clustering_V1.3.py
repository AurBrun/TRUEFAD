### Imports ###
import tkinter as tk
from tkinter import filedialog
import pandas as pd
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score
from sklearn.metrics import davies_bouldin_score
from sklearn.metrics import calinski_harabasz_score
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, GradientFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os

#GUI function
def select_file():
    filepath = filedialog.askopenfilename(title="Select The TRUEFAD Excel File", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    entry.delete(0, tk.END)
    entry.insert(0, filepath)

#Function that make fiber clusterization
def Cluster_image(init, df, Clusters):
    #Creating arrays for each fluorescence intensity corresponding to each fiber
    aT1_array = df.iloc[:, init].values
    bT1_array = aT1_array[~pd.isna(aT1_array)]
    T1_array = bT1_array[bT1_array != '']
    init += 1
    aT2A_array = df.iloc[:, init].values
    bT2A_array = aT2A_array[~pd.isna(aT2A_array)]
    T2A_array = bT2A_array[bT2A_array != '']
    init += 1
    #Trim the arrays to be even in case there is corruption of the native TRUEFAD excel file export, then concatenate them
    min_length = min(len(T1_array), len(T2A_array))
    T1 = T1_array[:min_length]
    T2A = T2A_array[:min_length]
    data = np.column_stack((T1, T2A))
    #Standardise the data and apply K-Mean clustering
    scaler = StandardScaler()
    scaled_data = scaler.fit_transform(data)
    kmeans = KMeans(n_clusters=Clusters, random_state=42)
    kmeans.fit(scaled_data)
    cluster_labels = kmeans.labels_
    sil = silhouette_score(scaled_data, cluster_labels)
    dav = davies_bouldin_score(scaled_data, cluster_labels)
    cal = calinski_harabasz_score(scaled_data, cluster_labels)
    ine = kmeans.inertia_
    Metrics = [sil, dav, cal, ine]
    return cluster_labels, Metrics

#Function to conditionally format cells on the exported excel
def format_excel(full_file_path):
    wb = openpyxl.load_workbook(full_file_path)
    sheet = wb['Metrics']
    min_row = 2
    max_row = sheet.max_row
    min_col = 1
    max_col = 3
    color_scale_rule_1 = ColorScaleRule(start_type='num', start_value=0, start_color='70FF0000', mid_type='num', mid_value=0.25, mid_color='70FFFF00', end_type='num', end_value=0.7, end_color='7000FF00')
    color_scale_rule_2 = ColorScaleRule(start_type='num', start_value=0.5, start_color='7000FF00', mid_type='num', mid_value=0.9, mid_color='70FFFF00', end_type='num', end_value=1.2, end_color='70FF0000')
    for col in range(min_col, max_col + 1):
        column_letter = get_column_letter(col)
        for row in range(min_row, max_row + 1):
            cell = sheet[f'{column_letter}{row}']
            if col == 1 : 
                sheet.conditional_formatting.add(f'{column_letter}{min_row}:{column_letter}{max_row}', color_scale_rule_1)
            elif col == 2 : 
                sheet.conditional_formatting.add(f'{column_letter}{min_row}:{column_letter}{max_row}', color_scale_rule_2) 
    #Adding information about the metrics
    strings = ['Silhouette Score: This metric measures how similar an object is to its own cluster (cohesion) compared to other clusters (separation). A score above 0.5 indicates a good clustering, a silhouette score below 0.25 indicates a bad clustering.'
                   , 'Davies-Bouldin Index: This index measures the average similarity between each cluster and its most similar cluster. Often used in order to evaluate the optimal number of clusters to use. Lower values under 1 indicate better separation.'
                   , 'Calinski-Harabasz Index: This index measures the ratio of between-cluster dispersion to within-cluster dispersion. Often used in order to evaluate the optimal number of clusters to use. Higher values indicate better separation.'
                   , 'Inertia: Inertia measures the sum of squared distances of samples to their closest cluster center. Lower values indicate better separation.']
    for i, value in enumerate(strings, start=2):
        sheet[f'H{i}'] = value
    cell = sheet['H2']
    bold_font = Font(bold=True)
    cell.font = bold_font
    wb.save(full_file_path)
    
#Open the file and works adaptatively on the dataframe
def read_excel():
    filepath = entry.get()
    RoundTHR = float(entry_1.get())
    AreaTHR = float(entry_2.get())
    Clusters = int(entry_3.get())
    if filepath:
        try:
            #Open the file a first time and store all images names as a list
            temp = pd.read_excel(filepath, header=None)
            temp_list = temp.iloc[0].tolist()
            Names_list = [x for x in temp_list if pd.notna(x)]
            #Intialize and reformat the Excel File for procedures
            df = pd.read_excel(filepath, header=1)
            df = df.replace(regex=',', value='')
            df = df.apply(pd.to_numeric, errors='coerce')
            #Detect blocks of images to process
            num_blocks = int((df.shape[1] + 1)/12)
            init = 6 #Main indexation parameter
            Metrics = []
            for block in range(num_blocks):                              
                cluster_labels, Metrics_block = Cluster_image(init, df, Clusters) #Run clusterization on T1, T2A signal for this block
                Metrics.append(Metrics_block)
                init +=5
                filler = df.shape[0] - len(cluster_labels) #Append the array with 0 to match the image that has the most fibers (number of rows of the dataframe)
                Clusterization = np.append(cluster_labels, [None] * filler)
                #Insert the clusterization column at the end of the block with the image name as the header
                df.insert(init, Names_list[block], Clusterization)
                #Save each of the column from the block as an array, then concatenate it as a 10-dimension array
                Count = df.iloc[:, init-11].values
                Area = df.iloc[:, init-10].values
                Perim = df.iloc[:, init-9].values
                Feret = df.iloc[:, init-8].values
                Circ = df.iloc[:, init-7].values
                Round = df.iloc[:, init-6].values
                T1 = df.iloc[:, init-5].values
                T2 = df.iloc[:, init-4].values
                Proba = df.iloc[:, init-3].values
                Type = df.iloc[:, init-2].values
                Olabel = df.iloc[:, init-1].values
                Label = df.iloc[:, init].values
                min_length = min(len(Count), len(Area), len(Perim), len(Feret), len(Circ), len(Round), len(T1), len(T2), len(Olabel), len(Type), len(Proba), len(Label))
                Count = Count[:min_length]
                Area = Area[:min_length]
                Perim = Perim[:min_length]
                Feret = Feret[:min_length]
                Circ = Circ[:min_length]
                Round = Round[:min_length]
                T1 = T1[:min_length]
                T2 = T2[:min_length]
                Proba = Proba[:min_length]
                Type = Type[:min_length]
                Olabel = Olabel[:min_length]
                Label = Label[:min_length]
                data = np.column_stack((Count, Area, Perim, Feret, Circ, Round, T1, T2, Proba, Type, Olabel, Label))
                #Iterate each row of the block and remove the indexed value if under the Area or Round threshold
                for row in data:
                    v0, v1, v2, v3, v4, v5, v6, v7, v8, v9, v10, v11 = row
                    if v1 < AreaTHR or v5 < RoundTHR:
                        for i in range(len(row)):
                            row[i] = np.nan
                #Replace the original columns with filtered data (blank space for removed ones)            
                df.iloc[:, init-11:init+1] = data[:, :init+1]
                print(Names_list[block], 'clusterization completed...')
                init +=8
            #Export the results and the metrics
            Header = ['Silhouette Score', 'Davies Bouldin Score', 'Calinski Harabasz Score', 'Inertia', 'Image']
            for i, row in enumerate(Metrics):
                row.append(Names_list[i])
            df_metrics = pd.DataFrame(Metrics, columns=Header)
            # Modify your code to use os.path.expanduser() to get the full file path
            file_path = '~/Desktop/TRUEFAD-Results.xlsx'
            full_file_path = os.path.expanduser(file_path)
            with pd.ExcelWriter(full_file_path) as writer:
                df_metrics.to_excel(writer, sheet_name='Metrics', index=False)
                temp.to_excel(writer, sheet_name='Raw', index=False, header=False)
                df.to_excel(writer, sheet_name='Cleaned', index=False)
            format_excel(full_file_path)
            print("JOB DONE!")
        except Exception as e:
            print("Error:", e)
    else:
        print("Please select a file first.")

#Create the main window
root = tk.Tk()
root.title("TRUEFAD-Histo-classification")

#Add numeric fields
label_1 = tk.Label(root, text="Remove fibers above this roundness:")
label_1.pack()
default_value_1 = tk.StringVar(value="0.25")  # Default value for field 1
entry_1 = tk.Entry(root, textvariable=default_value_1)
entry_1.pack()

label_2 = tk.Label(root, text="Remove fibers above this area (µm²):")
label_2.pack()
default_value_2 = tk.StringVar(value="300")  # Default value for field 2
entry_2 = tk.Entry(root, textvariable=default_value_2)
entry_2.pack()

label_3 = tk.Label(root, text="Number of clusters of interest:")
label_3.pack()
default_value_3 = tk.StringVar(value="3")  # Default value for field 2
entry_3 = tk.Entry(root, textvariable=default_value_3)
entry_3.pack()

#Add the path field
tk.Label(root, text="").pack()
label = tk.Label(root, text="Select Excel File:")
label.pack()
entry = tk.Entry(root, width=40)
entry.pack()

select_button = tk.Button(root, text="Select File", command=select_file)
select_button.pack()

read_button = tk.Button(root, text="Read Excel File", command=read_excel)
read_button.pack()

root.mainloop()

